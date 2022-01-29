#!/usr/bin/python3
#
# python3 -m pip install openpyxl - to install required module
'''
Convert test result in the RITdb to XLSX format
console window hidden on windows

Usage: RITdbVerify2xlsx.pyw [-s | --split] -i <input RITdb file> [-o <output XLSX file>]
'''

from fileinput import filename
import os
import sys
import getopt
import sqlite3 as sql
from openpyxl import Workbook
from openpyxl.styles import Font
import ctypes

# hides python terminal window when running from a Guru Agent on Windows
ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)

RED = Font(color="00FF0000")

PartInfoQuery = '''
SELECT  n2.value AS 'Part ID', n3.value AS 'Pass/Fail', n4.value AS 'Test Time', n5.value AS 'Cycle Time', n6.value AS 'Site', n1.value AS 'part result ID', n0.entityID AS 'entityID'
FROM ritdb1 n0
 JOIN ritdb1 n1 ON n0.entityID=n1.entityID AND n1.name='PART_RESULT_EVENT_ORDER'
 JOIN ritdb1 n2 ON n0.entityID=n2.entityID AND n2.name='PART_ID'
 JOIN ritdb1 n3 ON n0.entityID=n3.entityID AND n3.name='PF'
 JOIN ritdb1 n4 ON n0.entityID=n4.entityID AND n4.name='EVENT_TEST_TIME'
 JOIN ritdb1 n5 ON n0.entityID=n5.entityID AND n5.name='EVENT_CYCLE_TIME'
 JOIN ritdb1 n6 ON n0.entityID=n6.entityID AND n6.name='SITE_ID'
 {}
WHERE
 n0.value='PART_RESULT_EVENT'{}
ORDER BY n1.value ASC '''

Verify_ReslutInfoQuery = '''
SELECT  n1.value AS 'Result Number',  n2.value AS 'Result Name',  n4.value AS 'Units',  n8.value AS 'U Limit',  n9.value AS 'L Limit',  n3.value AS 'RESULT_ID'
FROM ritdb1 n0
 JOIN ritdb1 n1 ON n0.entityID=n1.entityID AND n1.indexID='0' AND n1.name='RESULT_NUMBER'
 JOIN ritdb1 n2 ON n0.entityID=n2.entityID AND n2.indexID='0' AND n2.name='RESULT_NAME'
 JOIN ritdb1 n3 ON n0.entityID=n3.entityID AND n3.indexID='0' AND n3.name='RESULT_ID'
 JOIN ritdb1 n4 ON n0.entityID=n4.entityID AND n4.indexID='0' AND n4.name='RESULT_UNITS'
 JOIN ritdb1 n6 ON n6.indexID='0' AND n6.name='ENTITY_TYPE' AND n6.value='RESULT_LIMIT_SET'
 JOIN ritdb1 n7 ON n6.entityID=n7.entityID AND 'SystemCheck'=n7.value AND n7.indexID='0' AND n7.name='LIMIT_SET_NAME'
 LEFT JOIN ritdb1 n8 ON n0.entityID=n8.indexID AND n6.entityID=n8.entityID AND n8.name='UL'
 LEFT JOIN ritdb1 n9 ON n0.entityID=n9.indexID AND n6.entityID=n9.entityID AND n9.name='LL'
WHERE
 n0.name='ENTITY_TYPE' AND n0.value='RESULT_INFO'
ORDER BY n3.value ASC'''

ResultsQuery = '''
SELECT  n0.value * n3.value AS 'Result',  n0.value2,  n1.value,  n2.value
FROM ritdb1 n0
 JOIN ritdb1 n1 ON n0.entityID=n1.entityID AND n1.name='PART_RESULT_EVENT_ORDER'
 JOIN ritdb1 n2 ON n0.indexID=n2.entityID AND n2.name='RESULT_ORDER'
 JOIN ritdb1 n3 ON n2.entityID=n3.entityID AND n3.name='RESULT_SCALE'
 {}
WHERE
 n0.name='R'{} '''

WaferEIDQuery = '''
SELECT n0.entityID AS 'WaferEID', n1.value AS 'WaferID'
FROM ritdb1 n0
 JOIN ritdb1 n1 ON n0.entityID=n1.entityID AND n1.name='SUBSTRATE_ID'
Where
 n0.value='SUBSTRATE_EVENT' '''

SubstrateEven_EID = '''JOIN ritdb1 n9 ON n0.entityID=n9.entityID AND n9.name='SUBSTRATE_EVENT_EID' '''


def ritdb2xlsx(dbCursor, xlFilename, modStr):
    failedCount = 0
    # query the top header
    if modStr:
        query = PartInfoQuery.format(SubstrateEven_EID, modStr)
    else:
        query = PartInfoQuery.format('', '')
    dbCursor.execute(query)
    topLabel = []
    for i in range(4):
        topLabel.append(dbCursor.description[i][0])

    # query the left table
    dbCursor.execute(Verify_ReslutInfoQuery)
    leftTable = dbCursor.fetchall()
    leftLabel = []
    for i in range(5):
        leftLabel.append(dbCursor.description[i][0])

    # query result data, sort by PART_RESULT_EVENT_ORDER then RESULT_ORDER
    if modStr:
        query = ResultsQuery.format(SubstrateEven_EID, modStr)
    else:
        query = ResultsQuery.format('', '')
    query = query + 'ORDER BY n2.value ASC'
    dbCursor.execute(query)

    # open Excel workbook
    wb = Workbook()
    ws1 = wb.active

    # workbook sheet title is limited to 31 chars on some versions
    ws1.title = os.path.basename(baseName)[:31]

    # write the top table to file
    ws1.cell(1, 6, os.path.basename(baseName))

    for i in range(5):
        colIdx = i + 1
        ws1.cell(1, colIdx, leftLabel[i])

    # Write one row at a time
    data = dbCursor.fetchone()      # get the first row of resultQuery data
    rowIdx = 2
    for ltRow in leftTable:
        for i in range(5):
            cell = ws1.cell(rowIdx, i + 1, ltRow[i])

        # leftTable.testEntityID = data.testEntityID
        while data != None and ltRow[5] == data[3]:
            cell = ws1.cell(rowIdx, 6, data[0])
            # Test UL and LL Limits
            if ltRow[3] and ltRow[3] < data[0]:
                cell.font = RED
                failedCount = failedCount + 1
            elif ltRow[4] and ltRow[4] > data[0]:
                cell.font = RED
                failedCount = failedCount + 1
            data = dbCursor.fetchone()
        rowIdx = rowIdx + 1

    # Set Verify PASS/FAIL Result
    if failedCount > 0:
        ws1.cell(rowIdx+1, 2, "FAIL")
        ws1.cell(rowIdx+1, 3, "Count")
        ws1.cell(rowIdx+1, 6).font = RED
        ws1.cell(rowIdx+1, 6, failedCount)
        ws1.cell(1, 6).font = RED  # Verify Name RED
    else:
        ws1.cell(rowIdx+1, 2, "PASS")

    wb.save(filename=xlFilename)
    print("SUCCESS converting to XLSX.", xlFilename)
### end of ritdb2xlsx ###


if __name__ == '__main__':

    usage = (
        "Usage: %s [-s | --split] -i <input RITdb file> [-o <output XLSX file>]" % (sys.argv[0]))

    dbFileName = ''
    xlsxFileName = ''
    pivot = False
    split = False

    argv = sys.argv[1:]
    try:
        opts, args = getopt.getopt(
            argv, "hpsi:o:", ["split", "ifile=", "ofile="])

    except getopt.GetoptError:
        print(usage)
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            print(usage)
            sys.exit()
        elif opt in ("-s", "--split"):
            split = True
        elif opt in ("-i", "--ifile"):
            dbFileName = arg.strip()
        elif opt in ("-o", "--ofile"):
            xlsxFileName = arg.strip()

    if not bool(dbFileName):    # dbFileName is empty
        print(usage)
        sys.exit(2)

    if bool(xlsxFileName):       # csvFileName is not empty
        baseName, ext = os.path.splitext(xlsxFileName)
        if ext:
            xlsxFileName = baseName + '{}' + ext
        else:
            xlsxFileName = baseName + '{}.xlsx'
    else:                       # csvFileName is empty
        baseName, ext = os.path.splitext(dbFileName)
        xlsxFileName = (baseName + '{}.xlsx')

    if not os.path.exists(dbFileName):
        print("RITdb file not found:", dbFileName)
        sys.exit(2)

    try:
        # Connect to database
        dbConn = sql.connect(dbFileName)
        dbCursor = dbConn.cursor()

        if split:
            dbCursor.execute(WaferEIDQuery)
            waferEID = dbCursor.fetchall()
            if waferEID:
                for eid in waferEID:
                    string = ' AND n9.value=' + str(eid[0])
                    newFileName = xlsxFileName.format(
                        f"{eid[1]:0>2}")  # pad one 0
                    ritdb2xlsx(dbCursor, newFileName, string)
                dbCursor.close()
                sys.exit(0)
            else:
                dbCursor.close()
                print("ERROR SPLIT - No wafer info found in RITdb:",
                      dbFileName, SubstrateEven_EID)
                sys.exit(2)

        newFileName = xlsxFileName.format('')
        ritdb2xlsx(dbCursor, newFileName, '')

    except sql.Error as error:
        print("Failed to read data from RITdb;", error)

    except Exception as e:
        print(e)

# Close database connection
    finally:
        dbConn.close()    # Close database connection
